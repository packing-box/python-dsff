# -*- coding: UTF-8 -*-
import builtins
import logging
import openpyxl.reader.excel as excelr
import types
from ast import literal_eval
from datetime import datetime
from functools import wraps
from getpass import getuser
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from zipfile import BadZipFile, ZipFile

from .__common__ import *
from .arff import *
from .csv import *
from .dataset import *


__all__ = ["DSFF"]


for name, etype in [("BadDsffFile", "OSError"), ("BadInputData", "ValueError"), ("EmptyDsffFile", "ValueError")]:
    if not hasattr(builtins, name):
        exec("class %s(%s): __module__ = 'builtins'" % (name, etype))
        setattr(builtins, name, locals()[name])


def _bind_from(dsff):
    def _adapt_name(f):
        def _wrapper(path, *args, **kwargs):
            r = f(dsff, path, *args, **kwargs)
            if dsff.name == INMEMORY:
                dsff.name = splitext(basename(path))[0]
            return r
        setattr(dsff, f.__name__, _wrapper)
        return _wrapper
    return _adapt_name


def _bind_to(dsff):
    def _is_empty(f):
        def _wrapper(*args, **kwargs):
            if len(dsff) == 0:
                raise EmptyDsffFile("No data")
            return f(dsff, *args, **kwargs)
        setattr(dsff, f.__name__, _wrapper)
        return _wrapper
    return _is_empty


class DSFF:
    """ DataSet File Format.
    
    Modes:          r   r+  w   w+
        --------------------------
        read        *   *       *
        write           *   *   *
        create              *   *
        truncate            *   *
    """
    def __init__(self, path=None, mode=None, logger=None):
        if mode is None:
            mode = "rw"[path in [None, INMEMORY]]
        if re.match(r"[rw]\+?$", mode) is None:
            raise ValueError("Mode should be one of: r, r+, w, w+")
        self.__change = False
        self.__logger = logger or logging.getLogger("DSFF")
        self.__name = None
        self.__path = path
        self.__mode = mode
        # depending on the mode, bind the necessary methods
        if mode in ["r+", "w", "w+"]:
            self.save = types.MethodType(lambda dsff: dsff._DSFF__save(), self)
        self.logger.debug("binding write methods")
        for name, obj in globals().items():
            if name.startswith("from_"):
                _bind_from(self)(obj)
        self.logger.debug("binding read methods")
        for name, obj in globals().items():
            if name.startswith("to_"):
                _bind_to(self)(obj)
        # perform checks
        if mode in ["r", "r+"]:
            if path is None:
                raise ValueError("No input path to a .dsff file provided")
            if path != INMEMORY and not isfile(path):
                raise FileNotFoundError("Input .dsff does not exist")            
        # if the target path exists and is a file, open it
        if mode in ["r", "r+"] and path != INMEMORY:
            # disable archive validation as it does not recognize '.dsff'
            tmp = excelr._validate_archive
            excelr._validate_archive = lambda f: ZipFile(f, 'r')
            try:
                self.__wb = load_workbook(path)
            except BadZipFile:
                raise BadDsffFile("File is not a DSFF file")
            finally:
                excelr._validate_archive = tmp
            # check that the file has only 2 worksheets: 'data' and 'features'
            if [ws._WorkbookChild__title for ws in self.__wb.worksheets] != ["data", "features"]:
                raise BadDsffFile("File is not a DSFF file")
            # check that the 'features' worksheet has 2 columns: 'name' and 'description'
            for headers in self.__wb['features'].rows:
                if len(headers) != 2 or headers[0].value != "name" or headers[1].value != "description":
                    raise BadDsffFile("The features worksheet does not comply with DSFF")
                break
            return
        # otherwise, create a new workbook with the default worksheets
        if isfile(self.path):
            remove(self.path)  # re-create
        self.__wb = Workbook()
        del self.__wb['Sheet']  # remove the default sheet
        for ws in ["data", "features"]:
            self.__wb.create_sheet(ws)
    
    def __enter__(self):
        return self
    
    def __exit__(self, *args):
        self.__save()
        self.close()
    
    def __getitem__(self, name):
        if name in ["data", "features"]:
            return self.__wb[name]
        # the common property 'description' is used to store the metadata of the dataset, hence 'description' can be
        #  used as a key in the metadata but not from the common properties
        try:
            if name != "description":
                return getattr(self.__wb.properties, name)
        except AttributeError:
            pass
        return self.metadata[name]
    
    def __len__(self):
        return len(self.data)
    
    def __setitem__(self, name, value):
        if name in ["data", "features"]:
            raise ValueError("'%s' is a name reserved for a worksheet" % name)
        # see the note from __getitem__ related to 'description'
        if hasattr(self.__wb.properties, name) and name != "description":
            setattr(self.__wb.properties, name, value)
        d = self.metadata
        d[name] = value
        self.__wb.properties.description = json.dumps(d)
        self.__change = True
    
    def __eval(self, v):
        try:
            return literal_eval(v)
        except (SyntaxError, ValueError):
            return v
    
    def __save(self):
        if self.mode == "r" or self.path == INMEMORY:
            return
        if self.__change:
            props = self.__wb.properties
            if props.creator is None or props.creator == "openpyxl":
                props.creator = getuser()
            props.title = self.name
            props.description = self.metadata
            if isfile(self.path) and self.mode.startswith("w"):
                remove(self.path)
            self.__wb.save(self.path)
            self.__change = False
    
    def close(self):
        self.__wb.close()
    
    def write(self, data=None, features=None, metadata=None, missing="?"):
        """ Write data and/or features and/or metadata to the workbook.
        
        :param data:     matrix of data (including headers) OR path to data.csv OR path to Dataset folder
        :param features: dictionary of features' names and descriptions OR path to features.json
        :param metadata: dictionary of dataset's metadata OR path to metadata.json
        """
        # get the cell coordinate from (X,Y) coordinates (e.g. (1,2) => "B1")
        coord = lambda x, y: ws.cell(x+1, y+1).coordinate
        # private function to auto-adjust column widths
        def autoadjust(ws):
            col_widths = []
            for row in ws.rows:
                if len(col_widths) == 0:
                    col_widths = len(row) * [0]
                for i, cell in enumerate(row):
                    col_widths[i] = max(col_widths[i], len(str(cell.value)))
            for i, w in enumerate(col_widths):
                ws.column_dimensions[get_column_letter(i+1)].width = w
        # if the first argument is a folder, assume it is a Dataset structure compliant with:
        #   name
        #    +-- data.csv
        #    +-- features.json
        #    +-- metadata.json
        if data is not None and not isinstance(data, (list, dict)) and isdir(expanduser(data)):
            self.__path, d = self.__path or basename(data), expanduser(data)
            data, features, metadata = join(d, "data.csv"), join(d, "features.json"), join(d, "metadata.json")
        # handle data first
        if data is not None:
            self.__logger.debug("writing data to DSFF...")
            ws, d = self.__wb['data'], data
            if not isinstance(d, list):
                if isfile(expanduser(d)) and splitext(d)[1] == ".csv":
                    with open(expanduser(d)) as f:
                        d = []
                        for row in csvmod.reader(f, delimiter=CSV_DELIMITER):
                            d.append(row)
                else:
                    raise BadInputData("'data' is not a list")
            for r, row in enumerate(d):
                for c, value in enumerate(row):
                    c = coord(r, c)
                    ws[c] = str({None: missing}.get(value, value))
                    if r == 0:
                        ws[c].alignment = Alignment(horizontal="center")
                        ws[c].font = Font(bold=True)
            autoadjust(ws)
            self.__change = True
        # then handle features dictionary
        if features is not None:
            self.__logger.debug("writing features to DSFF...")
            ws, headers, d = self.__wb['features'], ["name", "description"], features
            if not isinstance(d, dict):
                if isfile(expanduser(d)) and basename(d) == "features.json":
                    with open(expanduser(d)) as f:
                        d = json.load(f)
                else:
                    raise BadInputData("'features' is not a dictionary")
            try:
                for c, header in enumerate(headers):
                    c = coord(0, c)
                    ws[c] = header
                    ws[c].alignment = Alignment(horizontal="center")
                    ws[c].font = Font(bold=True)
                for r, pair in enumerate(d.items()):
                    ws[coord(r+1, 0)] = pair[0]
                    ws[coord(r+1, 1)] = pair[1]
                autoadjust(ws)
                self.__change = True
            except Exception as e:
                raise BadInputData("Unexpected error while parsing 'features' (%s)" % e)
        # finally handle metadata dictionary
        if metadata is not None:
            self.__logger.debug("writing metadata to DSFF...")
            d = metadata
            if not isinstance(d, dict):
                if isfile(expanduser(d)) and basename(d) == "metadata.json":
                    with open(expanduser(d)) as f:
                        d = json.load(f)
                else:
                    raise BadInputData("'metadata' is not a dictionary")
            try:
                self.__wb.properties.description = json.dumps(d)
            except Exception as e:
                raise BadInputData("Unexpected error while parsing 'metadata' (%s)" % e)
        self.__save()
    
    @property
    def data(self):
        return [[self.__eval(c.value) for c in cells] for cells in self.__wb['data'].rows]
    
    @property
    def features(self):
        return {cells[0].value: cells[1].value for i, cells in enumerate(self.__wb['features'].rows) if i > 0}
    
    @property
    def logger(self):
        return self.__logger
    
    @property
    def metadata(self):
        return json.loads((self.__wb.properties.description or "{}").replace("'", "\""))
    
    @property
    def mode(self):
        return self.__mode
    
    @property
    def name(self):
        return self.__name or self.__wb.properties.title or splitext(basename(self.path))[0]
    
    @name.setter
    def name(self, name):
        self.__name = name
    
    @property
    def path(self):
        p = self.__path or "undefined"
        if p != INMEMORY and not p.endswith(".dsff"):
            p += ".dsff"
        return p

